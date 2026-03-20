"""
Sizing & TCO Analysis — Fully Input-Driven Excel Calculator

Every parameter lives in a clearly marked, user-editable input cell.
Every computed cell is a formula referencing input cells only —
no hardcoded numeric constants inside formulas.
All memory units: MiB / GiB (binary, 1024-based).
xlsx internal formula format uses ',' for arguments (not ';').
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

# ============================================================
# Styles
# ============================================================
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
section_font = Font(bold=True, size=13, color="2F5496")
param_font = Font(bold=True, size=11)
value_font = Font(size=11)
input_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
input_font = Font(bold=True, size=11, color="9C6500")
result_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
result_font = Font(bold=True, size=11, color="2F5496")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
green_font = Font(bold=True, size=11, color="006100")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
red_font = Font(bold=True, size=11, color="9C0006")
total_font = Font(bold=True, size=13, color="FFFFFF")
total_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border


def style_row(ws, row, cols, font=value_font, fill=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True)
        if fill:
            cell.fill = fill


def style_input_row(ws, row, cols):
    """Style a row as user-editable input (yellow highlight)."""
    style_row(ws, row, cols, font=input_font, fill=input_fill)


# ============================================================
# Sheet 1: Hardware Sizing  (fully formula-driven)
# ============================================================
ws1 = wb.active
ws1.title = "1. Hardware Sizing"
ws1.column_dimensions["A"].width = 50
ws1.column_dimensions["B"].width = 24
ws1.column_dimensions["C"].width = 24
ws1.column_dimensions["D"].width = 48

# -- Title --
ws1.cell(1, 1, "VRAM SIZING CALCULATOR").font = Font(bold=True, size=14, color="2F5496")
ws1.cell(2, 1,
         "Yellow cells = user inputs (edit freely); "
         "blue cells = computed formulas").font = Font(
    italic=True, size=10, color="666666")

# -- Header row 3 --
for c, h in enumerate(["Parameter", "FP16", "INT4", "Note"], 1):
    ws1.cell(3, c, h)
style_header(ws1, 3, 4)

# ===================== ROW MAP =====================
# Row  4 : params (billions)             <- INPUT
# Row  5 : bytes per param               <- INPUT
# Row  6 : Model size (GiB)              <- FORMULA
# Row  7 : (blank)
# Row  8 : section label
# Row  9 : layers                        <- INPUT
# Row 10 : head_dim                      <- INPUT
# Row 11 : num_kv_heads                  <- INPUT
# Row 12 : seq_len                       <- INPUT
# Row 13 : KV cache dtype bytes          <- INPUT
# Row 14 : KV Cache per request (MiB)    <- FORMULA
# Row 15 : (blank)
# Row 16 : section label
# Row 17 : target RPM                    <- INPUT
# Row 18 : target RPS                    <- FORMULA
# Row 19 : avg latency (sec)             <- INPUT
# Row 20 : concurrency                   <- FORMULA
# Row 21 : KV Cache total (GiB)          <- FORMULA
# Row 22 : (blank)
# Row 23 : TOTAL VRAM (GiB)              <- FORMULA
# ===================================================

# ---------- Model weights (rows 4-6) ----------
ws1.cell(4, 1, "[INPUT] Parameters (billions)")
ws1.cell(4, 2, 70)
ws1.cell(4, 3, 70)
ws1.cell(4, 4, "Model parameter count")
style_input_row(ws1, 4, 4)

ws1.cell(5, 1, "[INPUT] Bytes per parameter (model weights)")
ws1.cell(5, 2, 2)
ws1.cell(5, 3, 0.5)
ws1.cell(5, 4, "FP16 = 2, INT8 = 1, INT4 = 0.5")
style_input_row(ws1, 5, 4)

ws1.cell(6, 1, "Model size (GiB)")
ws1.cell(6, 2, "=B4*1000000000*B5/(1024*1024*1024)")
ws1.cell(6, 3, "=C4*1000000000*C5/(1024*1024*1024)")
ws1.cell(6, 4, "params * 1e9 * bytes_per_param / 1024^3")
style_row(ws1, 6, 4, font=result_font, fill=result_fill)
ws1.cell(6, 2).number_format = "0.00"
ws1.cell(6, 3).number_format = "0.00"

# ---------- KV Cache per request (rows 8-14) ----------
ws1.cell(8, 1, "--- KV Cache (per request) ---")
style_row(ws1, 8, 4, font=param_font)

ws1.cell(9, 1, "[INPUT] Layers")
ws1.cell(9, 2, 80)
ws1.cell(9, 3, 80)
ws1.cell(9, 4, "Number of transformer layers")
style_input_row(ws1, 9, 4)

ws1.cell(10, 1, "[INPUT] head_dim")
ws1.cell(10, 2, 128)
ws1.cell(10, 3, 128)
ws1.cell(10, 4, "hidden_size / num_attention_heads")
style_input_row(ws1, 10, 4)

ws1.cell(11, 1, "[INPUT] num_kv_heads")
ws1.cell(11, 2, 8)
ws1.cell(11, 3, 8)
ws1.cell(11, 4, "KV-head groups (GQA); equals num_heads if MHA")
style_input_row(ws1, 11, 4)

ws1.cell(12, 1, "[INPUT] seq_len (tokens)")
ws1.cell(12, 2, 2048)
ws1.cell(12, 3, 2048)
ws1.cell(12, 4, "Average request length in tokens")
style_input_row(ws1, 12, 4)

# KV cache dtype
ws1.cell(13, 1, "[INPUT] KV cache dtype (bytes per element)")
ws1.cell(13, 2, 2)
ws1.cell(13, 3, 2)
ws1.cell(13, 4, "FP16 = 2, INT8 = 1, INT4 = 0.5")
style_input_row(ws1, 13, 4)

# KV cache per request (MiB) — FORMULA
ws1.cell(14, 1, "KV Cache per request (MiB)")
ws1.cell(14, 2, "=2*B9*B11*B10*B12*B13/(1024*1024)")
ws1.cell(14, 3, "=2*C9*C11*C10*C12*C13/(1024*1024)")
ws1.cell(14, 4, "2 * layers * kv_heads * head_dim * seq_len * dtype / 1024^2")
style_row(ws1, 14, 4, font=result_font, fill=result_fill)
ws1.cell(14, 2).number_format = "0.00"
ws1.cell(14, 3).number_format = "0.00"

# ---------- Throughput & concurrency (rows 16-21) ----------
ws1.cell(16, 1, "--- Throughput & Concurrency ---")
style_row(ws1, 16, 4, font=param_font)

ws1.cell(17, 1, "[INPUT] Target RPM")
ws1.cell(17, 2, 1000)
ws1.cell(17, 3, 1000)
ws1.cell(17, 4, "Throughput requirement")
style_input_row(ws1, 17, 4)

ws1.cell(18, 1, "Target RPS")
ws1.cell(18, 2, "=B17/60")
ws1.cell(18, 3, "=C17/60")
ws1.cell(18, 4, "RPM / 60")
style_row(ws1, 18, 4)
ws1.cell(18, 2).number_format = "0.00"
ws1.cell(18, 3).number_format = "0.00"

ws1.cell(19, 1, "[INPUT] Average latency (sec)")
ws1.cell(19, 2, 2)
ws1.cell(19, 3, 2)
ws1.cell(19, 4, "Expected inference latency per request")
style_input_row(ws1, 19, 4)

ws1.cell(20, 1, "Concurrency (simultaneous requests)")
ws1.cell(20, 2, "=ROUND(B18*B19,0)")
ws1.cell(20, 3, "=ROUND(C18*C19,0)")
ws1.cell(20, 4, "RPS * Latency (Little's Law)")
style_row(ws1, 20, 4)

# Total KV cache (GiB) — FORMULA
ws1.cell(21, 1, "KV Cache total (GiB)")
ws1.cell(21, 2, "=B14*B20/1024")
ws1.cell(21, 3, "=C14*C20/1024")
ws1.cell(21, 4, "KV_per_request_MiB * concurrency / 1024")
style_row(ws1, 21, 4, font=result_font, fill=result_fill)
ws1.cell(21, 2).number_format = "0.00"
ws1.cell(21, 3).number_format = "0.00"

# ---------- TOTAL VRAM (row 23) ----------
ws1.cell(23, 1, "TOTAL VRAM (GiB)")
ws1.cell(23, 2, "=B6+B21")
ws1.cell(23, 3, "=C6+C21")
ws1.cell(23, 4, "Model_GiB + KV_total_GiB")
style_row(ws1, 23, 4, font=total_font, fill=total_fill)
ws1.cell(23, 2).number_format = "0.00"
ws1.cell(23, 3).number_format = "0.00"

# ============ GPU catalogue (rows 26+) ============
R_GPU_TITLE = 26
ws1.cell(R_GPU_TITLE, 1, "GPU CATALOGUE").font = Font(
    bold=True, size=14, color="2F5496")
ws1.cell(R_GPU_TITLE + 1, 1,
         "Edit VRAM / TFLOPS to match your hardware; "
         "GPU requirement formulas reference these cells").font = Font(
    italic=True, size=10, color="666666")

R_GPU_HDR = 28
for c, h in enumerate(["GPU", "VRAM (GiB)", "FP16 TFLOPS", "Note"], 1):
    ws1.cell(R_GPU_HDR, c, h)
style_header(ws1, R_GPU_HDR, 4)

gpus = [
    ("NVIDIA A100 80 GiB (SXM)", 80, 312, "Best for large models"),
    ("NVIDIA A100 40 GiB (PCIe)", 40, 312, "VRAM-limited"),
    ("NVIDIA T4 16 GiB", 16, 65, "Budget inference"),
    ("NVIDIA L4 24 GiB", 24, 121, "Inference-optimized"),
    ("NVIDIA H100 80 GiB (SXM)", 80, 990, "Top-tier performance"),
]
for i, (name, vram, tflops, note) in enumerate(gpus, 1):
    r = R_GPU_HDR + i
    ws1.cell(r, 1, name)
    ws1.cell(r, 2, vram)
    ws1.cell(r, 3, tflops)
    ws1.cell(r, 4, note)
    style_input_row(ws1, r, 4)

# GPU catalogue row constants (for formula references)
GPU_A100_80 = R_GPU_HDR + 1  # 29
GPU_A100_40 = R_GPU_HDR + 2  # 30
GPU_T4 = R_GPU_HDR + 3       # 31
GPU_L4 = R_GPU_HDR + 4       # 32
# GPU_H100 = R_GPU_HDR + 5   # 33 (available for user configs)

# ============ GPU requirements table ============
R_REQ_TITLE = R_GPU_HDR + len(gpus) + 2  # 35
ws1.cell(R_REQ_TITLE, 1, "GPU REQUIREMENTS").font = Font(
    bold=True, size=14, color="2F5496")
ws1.cell(R_REQ_TITLE + 1, 1,
         "GPUs/replica = CEILING(Total VRAM / GPU VRAM from catalogue); "
         "Replicas = user input").font = Font(
    italic=True, size=10, color="666666")

R_REQ_HDR = R_REQ_TITLE + 2  # 37
ws1.column_dimensions["E"].width = 45
for c, h in enumerate(
        ["Configuration", "GPUs per replica", "[INPUT] Replicas",
         "Total GPUs", "Note"], 1):
    ws1.cell(R_REQ_HDR, c, h)
style_header(ws1, R_REQ_HDR, 5)

# CEILING formulas reference GPU catalogue VRAM cells (no hardcoded numbers)
R = R_REQ_HDR  # base row for data formulas
configs = [
    ("FP16 on A100-80 GiB",
     f"=CEILING(B23/B{GPU_A100_80},1)", 6, f"=B{R+1}*C{R+1}",
     "Tensor Parallel across GPUs"),
    ("FP16 on A100-40 GiB",
     f"=CEILING(B23/B{GPU_A100_40},1)", 6, f"=B{R+2}*C{R+2}",
     "Too many GPUs, inefficient"),
    ("FP16 on L4-24 GiB",
     f"=CEILING(B23/B{GPU_L4},1)", 6, f"=B{R+3}*C{R+3}",
     "Very complex configuration"),
    ("INT4 on A100-80 GiB",
     f"=CEILING(C23/B{GPU_A100_80},1)", 4, f"=B{R+4}*C{R+4}",
     "Fits on 1 GPU; 4 replicas enough"),
    ("INT4 on L4-24 GiB",
     f"=CEILING(C23/B{GPU_L4},1)", 5, f"=B{R+5}*C{R+5}",
     "2 GPUs in TP; 5 replicas"),
    ("INT4 on T4-16 GiB",
     f"=CEILING(C23/B{GPU_T4},1)", 6, f"=B{R+6}*C{R+6}",
     "Low throughput; not recommended"),
]
for i, (cfg, gpus_per, replicas, total_gpu, note) in enumerate(configs, 1):
    row = R_REQ_HDR + i
    ws1.cell(row, 1, cfg)
    ws1.cell(row, 2, gpus_per)
    ws1.cell(row, 3, replicas)
    ws1.cell(row, 4, total_gpu)
    ws1.cell(row, 5, note)
    if "INT4 on A100-80" in cfg:
        style_row(ws1, row, 5, font=green_font, fill=green_fill)
    elif "not recommended" in note:
        style_row(ws1, row, 5, font=red_font, fill=red_fill)
    else:
        style_row(ws1, row, 5)
    # Replica column (C) is user input — highlight it
    ws1.cell(row, 3).fill = input_fill
    ws1.cell(row, 3).font = input_font

# Row constants for cross-sheet references
# Row 38: FP16/A100-80, Row 41: INT4/A100-80, Row 42: INT4/L4, Row 43: INT4/T4
CFG_FP16_A100 = R_REQ_HDR + 1  # 38
CFG_INT4_A100 = R_REQ_HDR + 4  # 41
CFG_INT4_L4 = R_REQ_HDR + 5    # 42
CFG_INT4_T4 = R_REQ_HDR + 6    # 43


# ============================================================
# Sheet 2: Cloud Comparison
# ============================================================
ws2 = wb.create_sheet("2. Cloud Comparison")
ws2.column_dimensions["A"].width = 35
ws2.column_dimensions["B"].width = 22
ws2.column_dimensions["C"].width = 22
ws2.column_dimensions["D"].width = 22
ws2.column_dimensions["E"].width = 22
ws2.column_dimensions["F"].width = 35

ws2.cell(1, 1, "CLOUD COST COMPARISON").font = Font(
    bold=True, size=14, color="2F5496")
ws2.cell(2, 1,
         "Yellow cells = user inputs; blue cells = computed formulas").font = Font(
    italic=True, size=10, color="666666")

# -- Yandex Cloud --
ws2.cell(3, 1,
         "Yandex Cloud - GPU Instances (price per 1 GPU/hour)").font = section_font
for c, h in enumerate(
        ["GPU", "VRAM (GiB)", "Price (RUB/hr)", "Price (USD/hr)", "Note"], 1):
    ws2.cell(4, c, h)
style_header(ws2, 4, 5)

# USD price formulas reference the exchange-rate input cell ($B$9)
yc_gpus = [
    ("NVIDIA A100 80 GiB", 80, 538, "=C5/$B$9", "gpu-standard-v3"),
    ("NVIDIA T4 16 GiB",   16, 134, "=C6/$B$9", "g1/g2 series"),
    ("NVIDIA L4 24 GiB",   24, 201, "=C7/$B$9", "Available in some zones"),
]
for i, (gpu, vram, rub, usd, note) in enumerate(yc_gpus, 1):
    ws2.cell(4 + i, 1, gpu)
    ws2.cell(4 + i, 2, vram)
    ws2.cell(4 + i, 3, rub)
    ws2.cell(4 + i, 4, usd)
    ws2.cell(4 + i, 5, note)
    style_input_row(ws2, 4 + i, 5)
    # USD column is a formula, style as computed
    ws2.cell(4 + i, 4).number_format = "0.00"
    ws2.cell(4 + i, 4).fill = result_fill
    ws2.cell(4 + i, 4).font = result_font

# Row 8: (blank)
# Row 9: Exchange rate input
ws2.cell(9, 1, "[INPUT] Exchange rate (RUB per USD)")
ws2.cell(9, 2, 90)
style_input_row(ws2, 9, 2)

# -- Cloud.ru --
ws2.cell(10, 1,
         "Cloud.ru (SberCloud) - GPU Instances "
         "(price per 1 GPU/hour)").font = section_font
for c, h in enumerate(
        ["GPU", "VRAM (GiB)", "Price (RUB/hr)", "Price (USD/hr)", "Note"], 1):
    ws2.cell(11, c, h)
style_header(ws2, 11, 5)

cr_gpus = [
    ("NVIDIA A100 80 GiB", 80, 490, "=C12/$B$9", "ML Space / Evolution"),
    ("NVIDIA T4 16 GiB",   16, 120, "=C13/$B$9", "Basic GPU instance"),
    ("NVIDIA L4 24 GiB",   24, 185, "=C14/$B$9", "Inference-optimized"),
]
for i, (gpu, vram, rub, usd, note) in enumerate(cr_gpus, 1):
    ws2.cell(11 + i, 1, gpu)
    ws2.cell(11 + i, 2, vram)
    ws2.cell(11 + i, 3, rub)
    ws2.cell(11 + i, 4, usd)
    ws2.cell(11 + i, 5, note)
    style_input_row(ws2, 11 + i, 5)
    ws2.cell(11 + i, 4).number_format = "0.00"
    ws2.cell(11 + i, 4).fill = result_fill
    ws2.cell(11 + i, 4).font = result_font

# -- Monthly cost --
ws2.cell(17, 1, "MONTHLY COST BY CONFIGURATION").font = Font(
    bold=True, size=14, color="2F5496")
ws2.cell(18, 1, "[INPUT] Hours per month:")
ws2.cell(18, 2, 730)
style_input_row(ws2, 18, 2)

R_M = 20  # header row for monthly table
for c, h in enumerate(
        ["Configuration", "Num GPUs", "Yandex Cloud (RUB/mo)",
         "Cloud.ru (RUB/mo)", "Difference (RUB)", "Note"], 1):
    ws2.cell(R_M, c, h)
style_header(ws2, R_M, 6)

# GPU counts reference Sheet 1 total-GPUs column (D) — no hardcoded numbers.
# Pricing references this sheet's own input cells.
S1 = "'1. Hardware Sizing'"
monthly = [
    ("FP16, A100-80 GiB",
     f"={S1}!D{CFG_FP16_A100}",
     f"=B{R_M+1}*C5*$B$18", f"=B{R_M+1}*C12*$B$18",
     f"=C{R_M+1}-D{R_M+1}",
     "GPUs from Sheet 1 (edit replicas there)"),
    ("INT4, A100-80 GiB",
     f"={S1}!D{CFG_INT4_A100}",
     f"=B{R_M+2}*C5*$B$18", f"=B{R_M+2}*C12*$B$18",
     f"=C{R_M+2}-D{R_M+2}",
     "OPTIMUM - edit replicas on Sheet 1"),
    ("INT4, L4-24 GiB",
     f"={S1}!D{CFG_INT4_L4}",
     f"=B{R_M+3}*C7*$B$18", f"=B{R_M+3}*C14*$B$18",
     f"=C{R_M+3}-D{R_M+3}",
     "GPUs from Sheet 1 (edit replicas there)"),
    ("INT4, T4-16 GiB",
     f"={S1}!D{CFG_INT4_T4}",
     f"=B{R_M+4}*C6*$B$18", f"=B{R_M+4}*C13*$B$18",
     f"=C{R_M+4}-D{R_M+4}",
     "GPUs from Sheet 1 (edit replicas there)"),
]
for i, (cfg, ngpu, yc, cr, diff, note) in enumerate(monthly, 1):
    row = R_M + i
    ws2.cell(row, 1, cfg)
    ws2.cell(row, 2, ngpu)
    ws2.cell(row, 3, yc)
    ws2.cell(row, 4, cr)
    ws2.cell(row, 5, diff)
    ws2.cell(row, 6, note)
    for col in [2, 3, 4, 5]:
        ws2.cell(row, col).number_format = "#,##0"
    if "OPTIMUM" in note:
        style_row(ws2, row, 6, font=green_font, fill=green_fill)
    else:
        style_row(ws2, row, 6)


# ============================================================
# Sheet 3: Optimization
# ============================================================
ws3 = wb.create_sheet("3. Optimization")
ws3.column_dimensions["A"].width = 50
ws3.column_dimensions["B"].width = 22
ws3.column_dimensions["C"].width = 22
ws3.column_dimensions["D"].width = 40

ws3.cell(1, 1, "OPTIMIZATION: BATCHING & vLLM").font = Font(
    bold=True, size=14, color="2F5496")

ws3.cell(3, 1, "Optimization techniques").font = section_font
for c, h in enumerate(
        ["Technique", "Throughput gain", "Cost reduction", "Description"], 1):
    ws3.cell(4, c, h)
style_header(ws3, 4, 4)

opts = [
    ("Continuous Batching", "2-4x", "50-75%",
     "Batch requests dynamically; GPU util 30% -> 80%+"),
    ("vLLM (PagedAttention)", "2-5x", "50-80%",
     "Efficient KV Cache management; eliminates fragmentation"),
    ("vLLM + Continuous Batching", "3-6x", "65-85%",
     "Combined: PagedAttention + dynamic batching"),
    ("Speculative Decoding", "1.5-2.5x", "30-60%",
     "Draft model generates candidates, main verifies"),
    ("INT4 quantization (GPTQ/AWQ)", "1.5-2x", "50-70%",
     "4x less VRAM; fewer GPUs needed"),
]
for i, (tech, throughput, saving, desc) in enumerate(opts, 1):
    ws3.cell(4 + i, 1, tech)
    ws3.cell(4 + i, 2, throughput)
    ws3.cell(4 + i, 3, saving)
    ws3.cell(4 + i, 4, desc)
    style_row(ws3, 4 + i, 4)

# -- Cost forecast with optimization (formula-driven) --
OPT_TITLE = 12
ws3.cell(OPT_TITLE, 1,
         "COST FORECAST WITH OPTIMIZATION").font = Font(
    bold=True, size=14, color="2F5496")

# Reference cells — formulas pointing to Sheet 2 (single source of truth)
S2 = "'2. Cloud Comparison'"
ws3.cell(13, 1, "Cloud.ru A100 price (RUB/hr):")
ws3.cell(13, 2, f"={S2}!C12")
ws3.cell(13, 2).number_format = "#,##0"
ws3.cell(13, 3, "Hours/month:")
ws3.cell(13, 4, f"={S2}!B18")
style_row(ws3, 13, 4)
ws3.cell(13, 2).fill = result_fill
ws3.cell(13, 2).font = result_font
ws3.cell(13, 4).fill = result_fill
ws3.cell(13, 4).font = result_font

OPT_HDR = 15
for c, h in enumerate(
        ["Scenario", "[INPUT] Num GPUs", "Cost (RUB/mo)",
         "Savings vs Baseline"], 1):
    ws3.cell(OPT_HDR, c, h)
style_header(ws3, OPT_HDR, 4)

# Baseline GPU count references Sheet 1 INT4/A100-80 total GPUs.
# Optimization GPU counts are user inputs (hypothetical reductions).
scenarios = [
    ("Baseline (from Sheet 1: INT4, A100-80)",
     f"={S1}!D{CFG_INT4_A100}",
     f"=B{OPT_HDR+1}*$B$13*$D$13", None),
    ("[INPUT] + Continuous Batching", 3,
     f"=B{OPT_HDR+2}*$B$13*$D$13",
     f"=1-C{OPT_HDR+2}/C{OPT_HDR+1}"),
    ("[INPUT] + vLLM (PagedAttention)", 2,
     f"=B{OPT_HDR+3}*$B$13*$D$13",
     f"=1-C{OPT_HDR+3}/C{OPT_HDR+1}"),
    ("[INPUT] + vLLM + Batching + INT4", 2,
     f"=B{OPT_HDR+4}*$B$13*$D$13",
     f"=1-C{OPT_HDR+4}/C{OPT_HDR+1}"),
]
for i, (scenario, ngpu, cost, saving) in enumerate(scenarios, 1):
    row = OPT_HDR + i
    ws3.cell(row, 1, scenario)
    ws3.cell(row, 2, ngpu)
    ws3.cell(row, 3, cost)
    ws3.cell(row, 3).number_format = "#,##0"
    if saving is not None:
        ws3.cell(row, 4, saving)
        ws3.cell(row, 4).number_format = "0%"
    else:
        ws3.cell(row, 4, "---")
    if "vLLM + Batching" in scenario:
        style_row(ws3, row, 4, font=green_font, fill=green_fill)
    else:
        style_row(ws3, row, 4)
    # GPU count column is editable input (except baseline which is a formula)
    ws3.cell(row, 2).fill = input_fill
    ws3.cell(row, 2).font = input_font

# Baseline GPU cell is a formula, style it differently
ws3.cell(OPT_HDR + 1, 2).fill = result_fill
ws3.cell(OPT_HDR + 1, 2).font = result_font

# Key insight
R_INS = OPT_HDR + len(scenarios) + 2
ws3.cell(R_INS, 1, "KEY INSIGHT:").font = Font(
    bold=True, size=12, color="006100")
ws3.cell(R_INS + 1, 1,
         "Adjust GPU counts above to model your optimization scenario. "
         "Cost and savings update automatically.").font = value_font


# ============================================================
# Sheet 4: Recommendation  (dynamic dashboard)
# ============================================================
ws4 = wb.create_sheet("4. Recommendation")
ws4.column_dimensions["A"].width = 45
ws4.column_dimensions["B"].width = 55

ws4.cell(1, 1, "RECOMMENDATION DASHBOARD").font = Font(
    bold=True, size=14, color="2F5496")
ws4.cell(2, 1,
         "All numeric values update automatically from Sheets 1-3").font = Font(
    italic=True, size=10, color="666666")

for c, h in enumerate(["Parameter", "Value"], 1):
    ws4.cell(3, c, h)
style_header(ws4, 3, 2)

S3 = "'3. Optimization'"

r = 4

# --- VRAM Summary ---
ws4.cell(r, 1, "--- VRAM Summary ---")
style_row(ws4, r, 2, font=section_font)
r += 1

ws4.cell(r, 1, "Total VRAM: FP16 (GiB)")
ws4.cell(r, 2, f"={S1}!B23")
ws4.cell(r, 2).number_format = "0.00"
style_row(ws4, r, 2)
r += 1

ws4.cell(r, 1, "Total VRAM: INT4 (GiB)")
ws4.cell(r, 2, f"={S1}!C23")
ws4.cell(r, 2).number_format = "0.00"
style_row(ws4, r, 2)
r += 1

ws4.cell(r, 1, "VRAM reduction (FP16 / INT4)")
ws4.cell(r, 2, f"={S1}!B23/{S1}!C23")
ws4.cell(r, 2).number_format = '0.0"x"'
style_row(ws4, r, 2, font=result_font, fill=result_fill)
r += 2

# --- GPU Requirements (INT4, A100-80) ---
ws4.cell(r, 1, "--- GPU Requirements (INT4, A100-80) ---")
style_row(ws4, r, 2, font=section_font)
r += 1

ws4.cell(r, 1, "GPUs per replica")
ws4.cell(r, 2, f"={S1}!B{CFG_INT4_A100}")
style_row(ws4, r, 2)
r += 1

ws4.cell(r, 1, "Replicas")
ws4.cell(r, 2, f"={S1}!C{CFG_INT4_A100}")
style_row(ws4, r, 2)
r += 1

ws4.cell(r, 1, "Total GPUs")
ws4.cell(r, 2, f"={S1}!D{CFG_INT4_A100}")
style_row(ws4, r, 2, font=result_font, fill=result_fill)
r += 2

# --- Monthly Cost (Cloud.ru) ---
ws4.cell(r, 1, "--- Monthly Cost (Cloud.ru, RUB) ---")
style_row(ws4, r, 2, font=section_font)
r += 1

ws4.cell(r, 1, "Cost: no optimization")
ws4.cell(r, 2, f"={S3}!C{OPT_HDR+1}")
ws4.cell(r, 2).number_format = "#,##0"
style_row(ws4, r, 2)
r += 1

ws4.cell(r, 1, "Cost: with vLLM + Batching + INT4")
ws4.cell(r, 2, f"={S3}!C{OPT_HDR+4}")
ws4.cell(r, 2).number_format = "#,##0"
style_row(ws4, r, 2, font=green_font, fill=green_fill)
r += 1

ws4.cell(r, 1, "Savings (RUB/month)")
ws4.cell(r, 2, f"={S3}!C{OPT_HDR+1}-{S3}!C{OPT_HDR+4}")
ws4.cell(r, 2).number_format = "#,##0"
style_row(ws4, r, 2, font=green_font, fill=green_fill)
r += 1

ws4.cell(r, 1, "Savings %")
ws4.cell(r, 2, f"={S3}!D{OPT_HDR+4}")
ws4.cell(r, 2).number_format = "0%"
style_row(ws4, r, 2, font=green_font, fill=green_fill)
r += 2

# --- Qualitative Recommendations ---
ws4.cell(r, 1, "--- Qualitative Recommendations ---")
style_row(ws4, r, 2, font=section_font)
r += 1

qual_recs = [
    ("Inference Engine",
     "vLLM with PagedAttention + Continuous Batching"),
    ("Quantization",
     "INT4 (GPTQ or AWQ)"),
    ("Quality impact",
     "INT4 (AWQ/GPTQ) quality loss <1% on benchmarks"),
    ("Cloud provider",
     "Cloud.ru (SberCloud) - compare on Sheet 2"),
]
for param, value in qual_recs:
    ws4.cell(r, 1, param)
    ws4.cell(r, 2, value)
    style_row(ws4, r, 2)
    r += 1


# ============================================================
# Save
# ============================================================
output_path = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "Sizing_TCO_Calculator.xlsx",
)
wb.save(output_path)
print(f"Excel saved: {output_path}")
